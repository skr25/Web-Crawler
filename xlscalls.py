import openpyxl
import xlrd
from collections import OrderedDict
import json


def getLinkListFromWB(varWBPath):
	targetWorkbook = xlrd.open_workbook(varWBPath)
	sheet_names = targetWorkbook.sheet_names()
	desired_sheet = targetWorkbook.sheet_by_name(sheet_names[0])
	link_list = []
	for rownum in range(0,desired_sheet.nrows):
	    
	    row_values = desired_sheet.row_values(rownum)
	    link_list.append(row_values[0])
	# Serialize the list of dicts to JSON
	listOutput = link_list
	return listOutput

def getProductListFromWB(varWBPath,targetRow,targetCol):
	targetWorkbook = xlrd.open_workbook(varWBPath)
	sheet_names = targetWorkbook.sheet_names()
	desired_sheet = targetWorkbook.sheet_by_name(sheet_names[0])
	product_list = []
	for rownum in range(targetRow-1,desired_sheet.nrows):
	    
	    row_values = desired_sheet.row_values(rownum)
	    product_list.append(row_values[targetCol])
	# Serialize the list of dicts to JSON
	listOutput = product_list
	return listOutput
      
def createUserDebugWB(userList,miList,linkName):
	userWorkbook = openpyxl.Workbook()
	user_worksheet_first =  userWorkbook.active
	user_worksheet_first.title = "User"
	user_worksheet_first['A1'] = 'Feature'
	user_worksheet_first['B1'] = 'Customer-Testimonials'
	user_worksheet_first['C1'] = 'Case Study'
	user_worksheet_first['D1'] = 'Blog'
	user_worksheet_first['E1'] = 'Press'
	user_worksheet_first['F1'] = 'Total posts'
	user_worksheet_second = userWorkbook.create_sheet(title="Ml")
	user_worksheet_second['A1'] = 'Feature'
	user_worksheet_second['B1'] = 'Tag Name'
	user_worksheet_second['C1'] = 'Phrase'
	user_worksheet_second['D1'] = 'Address'
	user_worksheet_second['E1'] = 'Tag Count'
	user_worksheet_second['F1'] = 'Address_URL'
	user_worksheet_second['G1'] = 'URL_ADD'
	i=2
	for itemsInUser in userList:
	   user_worksheet_first['A%d' % i] = itemsInUser['feature']
	   user_worksheet_first['B%d' % i] = itemsInUser['customer-testimonial']
	   user_worksheet_first['C%d' % i] = itemsInUser['case-study']
	   user_worksheet_first['D%d' % i] = itemsInUser['blog']
	   user_worksheet_first['E%d' % i] = itemsInUser['press']
	   user_worksheet_first['F%d' % i] = itemsInUser['total-posts']
	   i += 1
	i=2   
	for itemsInMI in miList:
	   
	   user_worksheet_second['A%d' % i] = str(itemsInMI['feature'])
	   user_worksheet_second['B%d' % i] = str(itemsInMI['tagName'])
	   user_worksheet_second['C%d' % i] = str(itemsInMI['phrase'])
	   user_worksheet_second['D%d' % i] = str(itemsInMI['address'])
	   user_worksheet_second['E%d' % i] = str(itemsInMI['counts'])
	   user_worksheet_second['F%d' % i] = str(itemsInMI['addressUrl'])
	   user_worksheet_second['G%d' % i] = str(itemsInMI['urlAdd'])
	   i += 1
	   
	userWorkbook.save('Debug-%s.xlsx' % linkName )
	return True


def createUserWB(userList,miList,linkName):
	userWorkbook = openpyxl.Workbook()
	user_worksheet_first =  userWorkbook.active
	user_worksheet_first.title = "User"
	user_worksheet_first['A1'] = 'Feature'
	user_worksheet_first['B1'] = 'Customer-Testimonials'
	user_worksheet_first['C1'] = 'Case Study'
	user_worksheet_first['D1'] = 'Blog'
	user_worksheet_first['E1'] = 'Press'
	user_worksheet_first['F1'] = 'Total posts'
	user_worksheet_second = userWorkbook.create_sheet(title="Ml")
	user_worksheet_second['A1'] = 'Feature'
	user_worksheet_second['B1'] = 'Phrase'
	user_worksheet_second['C1'] = 'Address'
	user_worksheet_second['D1'] = 'Tag Count'
	user_worksheet_second['E1'] = 'Address_URL'
	user_worksheet_second['F1'] = 'URL_ADD'
	i=2
	for itemsInUser in userList:
	   user_worksheet_first['A%d' % i] = itemsInUser['feature']
	   user_worksheet_first['B%d' % i] = itemsInUser['customer-testimonial']
	   user_worksheet_first['C%d' % i] = itemsInUser['case-study']
	   user_worksheet_first['D%d' % i] = itemsInUser['blog']
	   user_worksheet_first['E%d' % i] = itemsInUser['press']
	   user_worksheet_first['F%d' % i] = itemsInUser['total-posts']
	   i += 1
	i=2 
	feature = []  
	strFeature = str(miList[0]['feature'])
	print(strFeature)
	feature.append(strFeature)
	
	phrase = []
	address = []
	counts = []
	addressUrl = []
	urlAdd = []
	for itemsInMI in miList:
            if str(itemsInMI['feature']) != strFeature:
                print("in if block..")
                print(str(itemsInMI['feature']))
                print("writing in row %d" % i)
                user_worksheet_second['A%d' % i] = feature[0]
                user_worksheet_second['B%d' % i] = str(phrase)
                user_worksheet_second['C%d' % i] = str(address)
                user_worksheet_second['D%d' % i] = str(counts)
                user_worksheet_second['E%d' % i] = str(addressUrl)
                user_worksheet_second['F%d' % i] = str(urlAdd)
                strFeature = str(itemsInMI['feature'])
                feature = []
                phrase = []
                address = []
                counts = []
                addressUrl = []
                urlAdd = []
                feature.append(strFeature)
                i += 1
            phrase.append(str(itemsInMI['phrase']))
            counts.append(str(itemsInMI['counts']))
            address.append(str(itemsInMI['address']))
            addressUrl.append(str(itemsInMI['addressUrl']))
            urlAdd.append(str(itemsInMI['urlAdd']))

	user_worksheet_second['A%d' % i] = feature[0]
	user_worksheet_second['B%d' % i] = str(phrase)
	user_worksheet_second['C%d' % i] = str(address)
	user_worksheet_second['D%d' % i] = str(counts)
	user_worksheet_second['E%d' % i] = str(addressUrl)
	user_worksheet_second['F%d' % i] = str(urlAdd)
	userWorkbook.save('%s.xlsx' % linkName )
	return True
        
